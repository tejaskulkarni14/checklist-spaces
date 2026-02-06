import streamlit as st
import openpyxl as op
from io import BytesIO

st.set_page_config(page_title="Excel Space Cleaner", page_icon="📊")

st.title("📊 Excel Space Cleaner")
st.write("Upload an Excel file to clean up spacing issues in a specific sheet")

# File upload
uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])

if uploaded_file is not None:
    # Load workbook to get sheet names
    workbook = op.load_workbook(uploaded_file)
    sheet_names = workbook.sheetnames
    
    # Display available sheets
    st.success(f"✅ File loaded successfully! Found {len(sheet_names)} sheet(s)")
    
    # Sheet selection
    selected_sheet = st.selectbox(
        "Select the sheet to process:",
        options=sheet_names,
        index=0
    )
    
    # Process button
    if st.button("🚀 Process File", type="primary"):
        with st.spinner("Processing your file..."):
            try:
                # Get the selected worksheet
                ws = workbook[selected_sheet]
                
                # Step 1: Add space after newlines if not present
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if '\n' in cell.value:
                                lines = cell.value.split("\n")
                                lines = [line + " " if not line.endswith(" ") else line for line in lines]
                                cell.value = "\n".join(lines)
                
                # Step 2: Replace double spaces with single spaces
                old_value = "  "
                new_value = " "
                
                while any(old_value in str(cell.value) for row in ws.iter_rows() for cell in row if cell.value):
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and old_value.lower() in str(cell.value).lower():
                                cell.value = str(cell.value).replace(old_value, new_value)
                
                # Save to BytesIO object
                output = BytesIO()
                workbook.save(output)
                output.seek(0)
                
                st.success("✨ Processing complete!")
                
                # Download button
                st.download_button(
                    label="📥 Download Processed File",
                    data=output,
                    file_name=f"{uploaded_file.name.rsplit('.', 1)[0]}_cleaned.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except Exception as e:
                st.error(f"❌ An error occurred: {str(e)}")
                
else:
    st.info("👆 Please upload an Excel file to get started")

# Instructions
with st.expander("ℹ️ What does this tool do?"):
    st.markdown("""
    This tool cleans up spacing issues in your Excel sheets:
    
    1. **Adds spaces after line breaks** - Ensures each line in multi-line cells ends with a space
    2. **Removes double spaces** - Replaces all instances of double spaces with single spaces
    
    **How to use:**
    - Upload your Excel file
    - Select the sheet you want to process
    - Click "Process File"
    - Download your cleaned file
    """)
